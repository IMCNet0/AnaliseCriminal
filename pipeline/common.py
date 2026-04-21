"""Utilidades compartilhadas pelos scripts de ingestão.

Objetivos:
- Ler arquivos .xlsx gigantes via openpyxl em modo read_only (não carrega tudo em RAM).
- Normalizar nomes de coluna (remove acento, upper, trim).
- Tratar "NULL" string, datas e coordenadas.
- Escrever Parquet particionado por ano/mês com dtypes enxutos.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterator, Iterable
import logging
import re

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from unidecode import unidecode

log = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent.parent
RAW_SSP = ROOT / "data" / "raw" / "ssp"
PROCESSED = ROOT / "data" / "processed"
AGGREGATES = ROOT / "data" / "aggregates"

NULL_TOKENS = {"NULL", "null", "Null", "", "NaN", "nan", "None"}


def norm_col(name: str) -> str:
    """Normaliza nome de coluna: remove acento, upper, substitui espaços/símbolos."""
    if name is None:
        return ""
    n = unidecode(str(name)).strip().upper()
    n = re.sub(r"[^A-Z0-9]+", "_", n).strip("_")
    return n


_DEFAULT_SKIP_SHEETS = (
    "METODOLOGIA",
    "DICIONARIO DE DADOS",
    "DICIONÁRIO DE DADOS",
    "CAMPOS DA TABELA_SPDADOS",
)


def _is_skip_sheet(name: str, skip_sheets: Iterable[str]) -> bool:
    """Case-insensitive comparison — tolera variações de capitalização/acento."""
    normalized = unidecode(name).strip().upper()
    for s in skip_sheets:
        if unidecode(s).strip().upper() == normalized:
            return True
    # Heurística extra: abas cujo nome começa com "CAMPOS" ou "METODOLOGIA"
    if normalized.startswith("CAMPOS") or normalized.startswith("METODOLOGIA"):
        return True
    return False


def read_xlsx_streaming(
    path: Path,
    sheet_name: str | None = None,
    chunksize: int = 50_000,
    skip_sheets: Iterable[str] = _DEFAULT_SKIP_SHEETS,
) -> Iterator[pd.DataFrame]:
    """Itera sobre chunks de um xlsx grande sem carregar tudo em RAM.

    Se sheet_name for None, percorre **todas** as abas de dados do arquivo
    (pula METODOLOGIA/DICIONARIO e variantes), preservando a ordem. Arquivos
    de anos completos da SSP costumam ter duas abas (ex.: "JAN-JUN_2022" e
    "JUL-DEZ_2022") porque o Excel trava em ~1M linhas por aba.
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        if sheet_name is not None:
            sheets_to_read = [sheet_name]
        else:
            sheets_to_read = [s for s in wb.sheetnames if not _is_skip_sheet(s, skip_sheets)]
        if not sheets_to_read:
            log.warning("Nenhuma aba de dados identificada em %s", path.name)
            return

        for sheet in sheets_to_read:
            ws = wb[sheet]
            rows_iter = ws.iter_rows(values_only=True)
            header_raw = next(rows_iter, None)
            if header_raw is None:
                log.warning("Aba vazia: %s", sheet)
                continue
            columns_raw = [norm_col(c) for c in header_raw]
            # Dedupe de colunas normalizadas: a SSP às vezes traz "Cidade" e
            # "CIDADE" (ou com espaço/acento) que colapsam após norm_col() e
            # quebram o pyarrow com "Duplicate column names found". Mantemos
            # a primeira ocorrência de cada nome.
            seen: set[str] = set()
            keep_idx: list[int] = []
            columns: list[str] = []
            for i, c in enumerate(columns_raw):
                if c in seen:
                    continue
                seen.add(c)
                keep_idx.append(i)
                columns.append(c)
            n_dupes = len(columns_raw) - len(columns)
            if n_dupes:
                dup_names = [c for c in columns_raw if columns_raw.count(c) > 1]
                log.warning(
                    "  · Aba '%s': %d coluna(s) duplicada(s) após normalização "
                    "(mantida primeira ocorrência): %s",
                    sheet, n_dupes, sorted(set(dup_names)),
                )
            n_cols = len(columns_raw)
            need_slice = n_dupes > 0
            log.info("  · Aba '%s' (%s colunas)", sheet, len(columns))

            buffer = []
            sheet_rows = 0
            for row in rows_iter:
                # Pula linhas 100% vazias (acontece em abas com linhas de rodapé)
                if all(v is None or v == "" for v in row):
                    continue
                # Pad/trim defensivo: algumas abas retornam rows com largura
                # inconsistente em relação ao header.
                if len(row) < n_cols:
                    row = row + (None,) * (n_cols - len(row))
                elif len(row) > n_cols:
                    row = row[:n_cols]
                if need_slice:
                    row = tuple(row[i] for i in keep_idx)
                buffer.append(row)
                if len(buffer) >= chunksize:
                    yield pd.DataFrame(buffer, columns=columns)
                    sheet_rows += len(buffer)
                    buffer.clear()
            if buffer:
                yield pd.DataFrame(buffer, columns=columns)
                sheet_rows += len(buffer)
            log.info("  · Aba '%s' concluída: %s linhas", sheet, f"{sheet_rows:,}")
    finally:
        wb.close()


def clean_null_strings(df: pd.DataFrame) -> pd.DataFrame:
    """Substitui 'NULL' (e variações) por NaN real."""
    obj_cols = df.select_dtypes(include="object").columns
    for c in obj_cols:
        df[c] = df[c].where(~df[c].isin(NULL_TOKENS), np.nan)
    return df


def parse_datetime_safe(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce")


def to_float_safe(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


_NP_INT_MAP = {"Int8": np.int8, "Int16": np.int16, "Int32": np.int32, "Int64": np.int64}


def to_nullable_int(series: pd.Series, dtype: str = "Int32") -> pd.Series:
    """Converte uma Series numérica para Int nullable (Int8/16/32/64) tolerando
    NaN/inf e valores fracionários.

    Por que não usar ``.astype('Int32')`` direto? No caminho interno do pandas
    ocorre um *safe cast* que compara ``values == casted_values`` para validar
    a conversão. Como ``NaN != NaN``, qualquer NaN na coluna dispara
    ``TypeError: cannot safely cast non-equivalent float64 to int32`` — mesmo
    após ``.round()`` e ``.replace([inf, -inf], NaN)``.

    Aqui construímos o ``IntegerArray`` manualmente: separamos a máscara de
    NaN, preenchemos com 0 antes do cast numérico (posições mascaradas nunca
    são lidas) e deixamos pd.NA apenas nas posições mascaradas.
    """
    if dtype not in _NP_INT_MAP:
        raise ValueError(f"dtype deve ser um de {list(_NP_INT_MAP)}; recebido: {dtype}")
    np_dtype = _NP_INT_MAP[dtype]

    s = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan)
    # Clippa valores fora do range do dtype para evitar overflow silencioso
    info = np.iinfo(np_dtype)
    s = s.where(s.isna() | ((s >= info.min) & (s <= info.max)))
    s = s.round()

    mask = s.isna().to_numpy()
    values = s.fillna(0).to_numpy().astype(np_dtype, copy=False)
    arr = pd.arrays.IntegerArray(values, mask=mask)
    return pd.Series(arr, index=series.index, name=series.name)


def valid_sp_bounds(lat: pd.Series, lon: pd.Series) -> pd.Series:
    """Máscara para coordenadas plausíveis dentro do estado de SP."""
    return (
        lat.between(-25.5, -19.5) & lon.between(-53.5, -44.0) & (lat != 0) & (lon != 0)
    )


def stringify_cols(df: pd.DataFrame, cols: Iterable[str]) -> pd.DataFrame:
    """Força colunas a tipo string (pd.StringDtype).

    Necessário antes de escrever em Parquet quando o Excel mistura int+str na
    mesma coluna (típico em NUM_BO, NUMERO_LOGRADOURO, CEP etc.). PyArrow
    rejeita colunas 'object' com tipos mistos.
    """
    for c in cols:
        if c in df.columns:
            df[c] = df[c].astype("string")
    return df


def stringify_object_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Converte TODAS as colunas 'object' restantes para string consistente.

    Fallback genérico: se alguma coluna escapou da lista explícita e ainda tem
    tipos mistos, isto evita o ArrowTypeError na escrita.
    """
    for c in df.select_dtypes(include="object").columns:
        df[c] = df[c].astype("string")
    return df


def optimize_dtypes(df: pd.DataFrame, categorical_cols: Iterable[str] = ()) -> pd.DataFrame:
    """No-op intencional para compatibilidade com os scripts de ingestão.

    Histórico de problemas evitados aqui:

    1. **Downcast numérico**: antes ``pd.to_numeric(df[c], downcast="integer")``
       escolhia o menor tipo que coubesse no chunk atual — um chunk com
       ``QUANTIDADE_OBJETO = [1,2,5]`` saía ``Int8``; outro com
       ``[168455, 42]`` saía ``Int32``. Cada partição Parquet escrevia com
       seu próprio schema → schema drift, leitura quebrava com
       ``Integer value X not in range``.
    2. **Conversão para `category`**: o pyarrow serializa ``category``
       como ``dictionary<indices=int8, values=string>`` e o tipo dos
       índices varia entre chunks (int8 se ≤128 valores únicos, int16 se
       mais). Cada partição saía com um tipo diferente de dicionário →
       drift idêntico ao do downcast, manifestando-se como
       ``Integer value 241 not in range: -128 to 127`` na unificação.

    Solução: manter ``string`` puro (feito por ``stringify_cols``) e
    ``Int*/Float64`` explícitos (feitos por ``to_nullable_int`` /
    ``to_float_safe``). O parquet resultante tem schema 100% estável
    entre partições. O overhead de memória é pequeno e o pyarrow ainda
    usa dicionário internamente (por row group) para compactar.
    """
    # Mantido por compatibilidade; não altera dtypes.
    _ = categorical_cols
    return df


def write_parquet_partitioned(df: pd.DataFrame, base_dir: Path, partition_cols: list[str]) -> None:
    """Anexa a um dataset parquet particionado por ano/mês."""
    base_dir.mkdir(parents=True, exist_ok=True)
    df.to_parquet(base_dir, index=False, partition_cols=partition_cols, engine="pyarrow")


def setup_logging() -> None:
    logging.basicConfig(
        format="%(asctime)s  %(levelname)-7s  %(message)s",
        level=logging.INFO,
        datefmt="%H:%M:%S",
    )
